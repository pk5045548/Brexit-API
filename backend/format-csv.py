import pandas as pd
import openpyxl
import os
import glob
import re
# Function to convert sheet data to CSV format
def sheet_to_csv(sheet, start_row):
    csv_data = []
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        csv_data.append(','.join(map(str, row)))
    return '\n'.join(csv_data)

# Function to write CSV data to a file
def write_csv_file(file_name, csv_data):
    with open(file_name, 'w', newline='', encoding='utf-8') as csv_file:
        csv_file.write(csv_data)

# Function to find the row where the data starts
def find_start_row(sheet, keyword):
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for cell in row:
            if cell is not None and keyword in str(cell):
                return row_idx
    return -1  # Return -1 if keyword is not found

# Function to get the appropriate keyword based on the sheet index
def get_keyword_for_sheet(index):
    if index == 6:
        return 'Nationality'
    elif index == 7:
        return 'Period'
    else:
        return 'Flow'

def remove_frozen_panes(sheet):
    sheet.freeze_panes = None

# Function to process the workbook and export specified sheets to CSV
def export_sheets_to_csv(workbook, output_folder):
    for index, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]
        
        # Remove frozen panes if they exist
        remove_frozen_panes(sheet)
        
        keyword = get_keyword_for_sheet(index)
        start_row = find_start_row(sheet, keyword)

        # Check if the sheet contains valid start_row
        if start_row == -1 or not (3 <= index <= 7):
            continue

        csv_data = sheet_to_csv(sheet, start_row)
        file_name = f"Migration_Data_{index - 2}.csv"
        file_path = os.path.join(output_folder, file_name)
        write_csv_file(file_path, csv_data)
        print(f"Exported {file_name}")

def change_csv_headers(input_file, output_file, header_mapping):
    # Read CSV into a DataFrame
    df = pd.read_csv(input_file)
    
    # Rename columns using the provided mapping
    df.rename(columns=header_mapping, inplace=True)
    
    # Replace blanks and 'None' with NaN
    df.replace('', pd.NA, inplace=True)
    df.replace('None', pd.NA, inplace=True)
    
    # Drop rows with any NaN values
    df.dropna(inplace=True)
    
    # Save the modified DataFrame to a new CSV file
    df.to_csv(output_file, index=False)

   
def format_migration_data(input_file, output_file, header_mapping):
    # Read CSV into a DataFrame
    df = pd.read_csv(input_file)
    
    # Rename columns using the provided mapping
    df.rename(columns=header_mapping, inplace=True)
    
    # Replace 'None' with NaN
    df.replace('None', pd.NA, inplace=True)
    
    # Drop columns that only contain NaN values
    df.dropna(axis=1, how='all', inplace=True)
    
    # Drop columns that have the header 'None' if they still exist
    df.drop(columns=[col for col in df.columns if 'None' in col], inplace=True)
    
    # Remove 'YE ' and ' P' characters from all columns
    df.replace(to_replace=[r'^YE\s+', r'\s+P$'], value='', regex=True, inplace=True)

    # Save the modified DataFrame to a new CSV file
    df.to_csv(output_file, index=False)

def replace_and_delete_rows(input_file, output_file, new_headers):
    try:
        # Read CSV into DataFrame
        df = pd.read_csv(input_file, skiprows=7)
        
        # Check if new_headers matches the number of columns

        if len(new_headers) != df.shape[1]:
            raise ValueError(f"Number of new headers ({len(new_headers)}) does not match number of columns in the CSV file ({df.shape[1]}).")
        
        # Insert new header row at the beginning
        df.columns = new_headers
        
        # Write modified DataFrame to new CSV file
        df.to_csv(output_file, index=False)
        
        print(f"Modified CSV file written to {output_file}")
        
    except Exception as e:
        print(f"Error: {e}")
        
def replace_and_delete_rows13(input_file, output_file, new_headers):
    try:
        # Read CSV into DataFrame
        df = pd.read_csv(input_file, skiprows=13)
        
        # Check if new_headers matches the number of columns
        if len(new_headers) != df.shape[1]:
            raise ValueError(f"Number of new headers ({len(new_headers)}) does not match number of columns in the CSV file ({df.shape[1]}).")
        
        # Insert new header row at the beginning
        df.columns = new_headers
        
        # Write modified DataFrame to new CSV file
        df.to_csv(output_file, index=False)
        
        print(f"Modified CSV file written to {output_file}")
        
    except Exception as e:
        print(f"Error: {e}")
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":

    # Load the workbook
    workbook = openpyxl.load_workbook('source-data/Migration-june23ltimaccessible.xlsx')
    output_folder = 'source-data'  # Specify the output folder

    # Export specified sheets to CSV
    export_sheets_to_csv(workbook, output_folder)

    # Example usage for HPI file:
    input_csv_hpi = 'source-data/UK-HPI-full-file-2024-02.csv'
    output_csv_hpi = 'source-data/UK-HPI-FIXED.csv'

    headers_mapping_hpi = {
        'Date' : 'Date1',
        'Index': 'HPIINDEX',
        '1m%Change': 'OneMonthChange',
        '12m%Change': 'TwelveMonthChange',
        'Detached1m%Change': 'DetachedOneMonthChange',
        'Detached12m%Change': 'DetachedTwelveMonthChange',
        'SemiDetached1m%Change': 'SemiDetachedOneMonthChange',
        'SemiDetached12m%Change': 'SemiDetachedTwelveMonthChange',
        'Terraced1m%Change': 'TerracedOneMonthChange',
        'Terraced12m%Change': 'TerracedTwelveMonthChange',
        'Flat1m%Change': 'FlatOneMonthChange',
        'Flat12m%Change': 'FlatTwelveMonthChange',
        'Cash1m%Change': 'CashOneMonthChange',
        'Cash12m%Change': 'CashTwelveMonthChange',
        'Mortgage1m%Change': 'MortgageOneMonthChange',
        'Mortgage12m%Change': 'MortgageTwelveMonthChange',
        'FTB1m%Change': 'FTBOneMonthChange',
        'FTB12m%Change': 'FTBtwelveMonthChange',
        'FOO1m%Change': 'FOOOneMonthChange',
        'FOO12m%Change': 'FOOTwelveMonthChange',
        'New1m%Change': 'NewOneMonthChange',
        'New12m%Change': 'NewTwelveMonthChange',
        'Old1m%Change': 'OldOneMonthChange',
        'Old12m%Change': 'OldTwelveMonthChange'
    }

    change_csv_headers(input_csv_hpi, output_csv_hpi, headers_mapping_hpi)
    
    input_csv_migration = 'source-data/Migration_Data_1.csv'
    output_csv_migration = 'source-data/Migration_Data_1_FIXED.csv'
    headers_mapping_migration = {
        # Define your header mapping dictionary for the Migration data here
        'Non-EU': 'NonEU',
        'All Nationalities': 'AllNationalities'
    }
    format_migration_data(input_csv_migration, output_csv_migration, headers_mapping_migration)

    input_csv_migration = 'source-data/Migration_Data_2.csv'
    output_csv_migration = 'source-data/Migration_Data_2_FIXED.csv'
    headers_mapping_migration = {
        # Define your header mapping dictionary for the Migration data here
        'Lower bound': 'Lowerbound',
        '25% limit' : 'TwentyFivePercentLimit',
        '75% limit' : 'SeventyFivePercentLimit',
        'Upper bound': 'Upperbound',       
    }
    format_migration_data(input_csv_migration, output_csv_migration, headers_mapping_migration)

    input_csv_migration = 'source-data/Migration_Data_3.csv'
    output_csv_migration = 'source-data/Migration_Data_3_MERGED.csv'
    new_headers = ['Flow', 'Period', 'AllReasons', 'Work', 'WorkDependant', 'Study', 'StudyDependant', 'Family', 'Other', 'HumanitarianBNO', 'HumanitarianResettlement', 'HumanitrianUkraine', 'Asylum', 'None', 'None', 'None', 'None']

    replace_and_delete_rows(input_csv_migration, output_csv_migration, new_headers)

    input_csv_migration = 'source-data/Migration_Data_3_MERGED.csv'
    output_csv_migration = 'source-data/Migration_Data_3_FIXED.csv'
    headers_mapping_migration = {
        # Define your header mapping dictionary for the Migration data here
        'Other ': 'Other',    
        'HumanitrianUkraine':'HumanitarianUkraine'
    }

    format_migration_data(input_csv_migration, output_csv_migration, headers_mapping_migration)

    input_csv_migration = 'source-data/Migration_Data_4.csv'
    output_csv_migration = 'source-data/Migration_Data_4_FIXED.csv'
    headers_mapping_migration = {
        # Define your header mapping dictionary for the Migration data here
        'All Reasons': 'AllReasons',    
    }

    format_migration_data(input_csv_migration, output_csv_migration, headers_mapping_migration)


    input_csv_migration = 'source-data/Migration_Data_5.csv'
    output_csv_migration = 'source-data/Migration_Data_5_MERGED.csv'
    new_headers = [
    "Period", "Publication", "ImmigrationAll", "ImmigrationBritish", "ImmigrationEU", 
    "ImmigrationNonEU", "EmigrationAll", "EmigrationBritish", "EmigrationEU", 
    "EmigrationNonEU", "NetAll", "NetBritish", "NetEU", "NetNonEU"
    
    ]
    replace_and_delete_rows13(input_csv_migration, output_csv_migration, new_headers)


    input_csv_migration = 'source-data/Migration_Data_5_MERGED.csv'
    output_csv_migration = 'source-data/Migration_Data_5_FIXED.csv'
    headers_mapping_migration = {
        # Define your header mapping dictionary for the Migration data here
        'Other ': 'Other',    
    }

    format_migration_data(input_csv_migration, output_csv_migration, headers_mapping_migration)